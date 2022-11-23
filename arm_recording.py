import os.path
import numpy as np
# import matplotlib.pyplot as plt
import serial
from timeit import default_timer as timer
import csv
import sys
import xlsxwriter
import winsound
import keyboard
import openpyxl
# from sim import Plot
from matplotlib import pyplot as plt
from plotter import Plot


class ArmRecording:

    def __init__(self, time_step, time_recording, num_sensors, port, coeffs_file, skip_serial_input=False):
        self.plot = Plot(time_recording)
        plt.pause(0.05)

        self.num_sensors = num_sensors  # please enter the number of sensors in the array. Make sure to also enter
        # the same number in the arduino array
        self.time_step = time_step  # 0.0500 (=50ms) #time interval in seconds to read data from the ARDUINO(values
        # has to correspond with the one in the ARDUINO code)
        self.time_recording = time_recording  # recording time in seconds
        self.num_measurements = int(
            time_recording / time_step)  # 100 #number of expected measurements. If this number is exceeded, the data
        # are saved and the program is closed

        self.raw_data = np.zeros((self.num_measurements, self.num_sensors + 3))
        self.lin_data = np.zeros((self.num_measurements, self.num_sensors + 3))

        self.I2C_address = np.array(
            [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32,
             33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59,
             60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 0x0F, 77, 78, 79, 0x0A, 0x0B, 0x0C, 0x0D,
             0x0E])

        self.reading_count = 0  # the first index contains the number of recorded data packages at the current
        # iteration/time; the second index contains the number of recorded data sets that got saved already (during a
        # previous intermediate saving action), the third number is the number of recorded data sets of theprevious
        # time step

        self.init_coeffs(coeffs_file)

        if skip_serial_input:
            return
        self.ser = serial.Serial(port, 57600)
        self.ser.flushInput()


    def init_coeffs(self, file):

        coeff_data = list(csv.reader(open(file, newline=''), delimiter=','))

        self.coeff_ref = np.array([float(coeff_data[1][0]), float(coeff_data[1][1])])

        self.x_coeffs =  {
            "up": np.asarray(coeff_data[2][:], dtype=float),
            "down": np.asarray(coeff_data[3][:], dtype=float)
        }

        self.y_coeffs = {
            "up": np.asarray(coeff_data[4:(4 + self.num_sensors)][:], dtype=float),
            "down": np.asarray(coeff_data[(4 + self.num_sensors):(4 + self.num_sensors + self.num_sensors)][:],
                             dtype=float)
        }

        self.num_nodes = len(self.x_coeffs["up"])

    def record(self, linearize_realtime=False):

        self.ser.flush()
        self.ser.write(b'R\n')  # Tell Arduino that we DON'T want to send the transfer functions but record directly
        recording_confirm = False
        # wait till arduino gets the message that recording started
        while not recording_confirm:
            ser_input = self.ser.readline()[:-2].decode('ascii')  # the last bit gets rid of the new-line chars

            if ser_input[0] == "R":
                print("Data recording confirmed")
                recording_confirm = True

        self.ser.flush()

        # START OF RECORDING
        total_time = 0
        previous_time = timer()
        while self.reading_count < self.num_measurements:  # read the specified amount of data steps from sensor array
            print(self.reading_count)
            if keyboard.is_pressed("q"):
                print("q pressed, ending recording")
                break

            while True:
                current_time = timer()
                time_interval = current_time - previous_time

                #####READ SERIAL INPUT #######################################################
                ##Note: when we start reading, we do not wait the time interval
                if time_interval >= self.time_step or total_time == 0:
                    previous_time = current_time
                    total_time = float("{:.2f}".format(total_time + time_interval))  # limit time_interval to two digits
                    ser_input = self.ser.readline()[:-2]  # convert binary to string array
                    break
            ser_input = str(ser_input.decode("ascii")).split(",")
            # ######################Parse Serial INPUT ############################################### TODO: Why when
            #  accessing in ser_input, we use index % ser_size (Why the modulo size) TODO: directly append the whole
            #   array to self.raw_data array, instead of predefine an empty array and then one by one filling it
            ser_size = len(ser_input)
            self.raw_data[self.reading_count][0] = total_time
            self.lin_data[self.reading_count][0] = total_time

            for i in range(self.num_sensors):
                sensor_index = i + 1
                datapoint = float(ser_input[i % ser_size])
                self.raw_data[self.reading_count][sensor_index] = datapoint

                self.detect_sensor_errors(datapoint, sensor_index)

            # write down arm position in degrees
            arm_angle = ser_input[self.num_sensors % ser_size]
            self.raw_data[self.reading_count][self.num_sensors + 1] = arm_angle
            self.lin_data[self.reading_count][self.num_sensors + 1] = arm_angle
            # write down orthosis position in degrees
            orthosis_angle = ser_input[self.num_sensors + 1 % ser_size]
            self.raw_data[self.reading_count][self.num_sensors + 2] = orthosis_angle
            self.lin_data[self.reading_count][self.num_sensors + 2] = orthosis_angle

            # optional, can linearize later
            if linearize_realtime:
                self.linearize_data_unit(self.raw_data[self.reading_count], self.reading_count)

            current_data = self.raw_data[self.reading_count] if not linearize_realtime else self.lin_data[self.reading_count]
            # print(self.reading_count)
            # FIXME plotting here
            self.plot.update_pressure_plot(current_data[1:49], current_data[49:71], current_data[71:-2])
            self.plot.update_angle_plot(current_data[0], current_data[-2], current_data[-1])
            # print("angle: ", current_data[-2])
            plt.pause(0.001)

            self.detect_serial_input_errors(ser_size)
            self.reading_count += 1

        if not linearize_realtime:
            self.linearize_all(self.raw_data)

    def linearize_from_file(self, file):

        raw_f = openpyxl.load_workbook(file).active
        # raw_data = np.array(list(csv.reader(raw_f, delimiter=',')))

        self.raw_data = np.zeros((raw_f.max_column, raw_f.max_row - 2))

        for i in range(int(raw_f.max_column)):
            self.raw_data[i][0] = raw_f.cell(2, i + 1).value
            self.lin_data[i][0] = raw_f.cell(2, i + 1).value

        for i in range(3, int(raw_f.max_row)):

            for j in range(int(raw_f.max_column)):
                self.raw_data[j][i - 2] = raw_f.cell(i + 1, j + 1).value

        self.linearize_all(self.raw_data)

    def linearize_all(self, raw_data):

        for i in range(len(raw_data)):
            self.linearize_data_unit(raw_data[i], i)

    def linearize_data_unit(self, raw_data_unit, unit_index):
        trailing_mean_length = 3
        countr = 0
        countf = 0

        # if condition checking for falling or rising edge
        for i in range(self.num_sensors):
            sensor_index = i + 1

            trailing_mean = self.trailing_mean(unit_index, sensor_index, trailing_mean_length)

            x_coeffs = None
            y_coeffs = None
            min_n = None
            max_n = None
            ##########################################################Check for Rising Signal######################################################
            # if prev. reading smaller than the current reading, then we have a rising signal
            if raw_data_unit[sensor_index] >= trailing_mean:
                # print('---Rising Pressure---')
                countr += 1
                x_coeffs = self.x_coeffs["up"]
                y_coeffs = self.y_coeffs["up"]
                min_n, max_n = self.find_coefficient_node(raw_data_unit[sensor_index], y_coeffs[sensor_index - 1][:])

            # pressure is falling - use falling pressure coefficients
            else:
                countf += 1
                x_coeffs = self.x_coeffs["down"]
                y_coeffs = self.y_coeffs["down"]
                min_n, max_n = self.find_coefficient_node(raw_data_unit[sensor_index], y_coeffs[sensor_index - 1], reverse=True)



            a = (self.coeff_ref[0] * x_coeffs[min_n] + self.coeff_ref[1])
            b = (((self.coeff_ref[0] * x_coeffs[max_n] + self.coeff_ref[1]) - (
                    self.coeff_ref[0] * x_coeffs[min_n] + self.coeff_ref[1])) / (
                         y_coeffs[sensor_index - 1][max_n] - y_coeffs[sensor_index - 1][min_n]))
            c = (raw_data_unit[sensor_index] - y_coeffs[sensor_index - 1][min_n])

            self.lin_data[unit_index][sensor_index] = a + b * c

    # returns the node with the closes value to the sensor value
    # reverse = True -> range is sorted descending
    def find_coefficient_node(self, sensor_value, sensor_node_range, reverse=False):

        min_n = 0 if not reverse else self.num_nodes - 1

        max_n = self.num_nodes - 1 if not reverse else 0

        while abs(max_n - min_n) > 1:

            mid_n = (max_n + min_n) // 2

            if sensor_value < sensor_node_range[mid_n]:
                max_n = mid_n
            else:
                min_n = mid_n

        return (min_n, max_n) if not reverse else (max_n, min_n)

    # compute mean of trailing_len amount of sensor values of a given sensor with index data_index
    def trailing_mean(self, unit_index, data_index, trail_len):
        if unit_index - trail_len < 0: return self.raw_data[unit_index][data_index]

        begin_index = max(0, unit_index - trail_len)

        return self.raw_data[begin_index:unit_index, data_index].mean()

    # TODO: why index - 1 on third if condition, CHECK!!!
    def detect_sensor_errors(self, datapoint, index):

        if datapoint > 511:
            print('Warning!!!!!! Entering overpressure range of sensor number 0x', self.I2C_address[index])
            winsound.Beep(4000, 2)
        if datapoint > 760:  # sensor range limit of S10N with a scaling factor of 2.4 is 767 corresponding to
            # 15N. This is no the max limit the sensor can handle
            print('Warning!!!!!! Absolut sensor range limit of sensor number 0x', self.I2C_address[index])
            winsound.Beep(8000, 2)
        if datapoint == -250 or datapoint == -251:
            print('Warning!!!!!! Faulty breakout board of sensor 0x', self.I2C_address[index])
        if datapoint == 772:
            print('Warning!!!!!! Faulty connection of sensor 0x', self.I2C_address[index])

    def detect_serial_input_errors(self, ser_size):

        if ser_size < self.num_sensors + 2:
            print('WARNING!! LESS SENSOR READINGS RECEIVED THAN in "num_sensors" SPECIFIED')
        elif ser_size > self.num_sensors + 2:
            print('WARNING!! MORE SENSOR READINGS RECEIVED THAN in "num_sensors" SPECIFIED')

    def save_data(self, folder, file_name):

        print('Saving Data in .CSV File')

        # calculating transposing so that data can be saved in same format as calibration data from Matlab
        raw_transpose = self.raw_data.transpose()
        lin_transpose = self.lin_data.transpose()

        print('FINAL DATA SAVING - FIRST SAVE ACTION SINCE RECORDING START')
        # writing data to .csv file:
        # Save raw data to xlsx:
        file_raw = os.path.join(folder, file_name + '_RAW.xlsx')
        workbook = xlsxwriter.Workbook(file_raw)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, file_name)

        for col_num, data in enumerate(raw_transpose[0]):
            worksheet.write(1, col_num, data)

        for i in range(self.num_sensors):
            index = i + 1
            for col_num, data in enumerate(raw_transpose[index]):
                worksheet.write(index + 2, col_num, data)

        for i in range(self.num_sensors + 1, self.num_sensors + 3):
            for col_num, data in enumerate(raw_transpose[i][:]):
                worksheet.write(i + 3, col_num, data)

        workbook.close()


        file_lin = os.path.join(folder, file_name + '_LIN.xlsx')
        workbook = xlsxwriter.Workbook(file_lin)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, file_name)

        for col_num, data in enumerate(lin_transpose[0]):
            worksheet.write(1, col_num, data)

        for i in range(self.num_sensors):
            index = i + 1
            for col_num, data in enumerate(lin_transpose[index]):
                worksheet.write(index + 2, col_num, data)

        for i in range(self.num_sensors + 1, self.num_sensors + 3):
            for col_num, data in enumerate(raw_transpose[i][:]):
                worksheet.write(i + 3, col_num, data)
        workbook.close()
        sys.exit("Data saved - Program closed")


def run_recording(path, file_name, time_recording):
    recording = ArmRecording(0.5, time_recording, 80, 'COM3', 'coefficients_S10N_14nodes_final_80Sensors.csv')
    recording.save_data(path, file_name)

