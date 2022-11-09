import numpy as np
import openpyxl


class Data:

    def __init__(self, path):

        # Arrays of pressure input from patches
        self.pressure_back_bottom, self.pressure_front, self.pressure_back_top = None, None, None

        # Arrays of angle values input from arm and orthosis
        self.angle_orthosis, self.angle_arm = None, None
        self.time_step_recording, self.time = None, None
        self.num_sensors, self.num_measurements = None, None
        # Excel sheet with data from recording
        self.sheet = None
        # Maximum pressure registered from all patches
        self.max_pressure = 0

        # Path to the xls data file from recording
        self.path = path
        self.init()

    def init(self):
        """
        Read the data from the recording
        """
        self.sheet = openpyxl.load_workbook(self.path).active

        self.num_measurements = self.sheet.max_column
        self.num_sensors = self.sheet.max_row - 6 if self.path[-8:-5] == "LIN" else self.sheet.max_row - 3
        self.time = self.read_time()
        self.time_step_recording = (self.time[1] - self.time[0]) * 1000

        self.angle_arm, self.angle_orthosis = self.read_angles()
        self.pressure_front = self.read_pressure_front()
        self.pressure_back_top = self.read_pressure_back_top()
        self.pressure_back_bottom = self.read_pressure_back_bottom()

    def read_time(self):
        """
        :return: time data from Excel file
        """
        _time = np.zeros(self.num_measurements)
        for i in range(1, self.num_measurements):
            # Time in secs
            _time[i] = self.sheet.cell(column=i, row=2).value
        return _time

    def read_angles(self):
        """
        :return: angle arm data and orthosis angle data from Excel file
        """
        _angle_arm = np.zeros(self.num_measurements)
        _angle_orthosis = np.zeros(self.num_measurements)

        for i in range(1, self.num_measurements):
            # Read arm angle
            _angle_orthosis[i] = self.sheet.cell(column=i, row=self.num_sensors + 6).value
            if _angle_orthosis[i] > 1000 or _angle_orthosis[i] < 0:
                _angle_orthosis[i] = 0

            # Read orthosis angle
            _angle_arm[i] = self.sheet.cell(column=i, row=self.num_sensors + 5).value
            if _angle_arm[i] > 1000 or _angle_arm[i] < 0:
                _angle_arm[i] = 0

        return _angle_orthosis, _angle_arm

    def read_pressure_front(self):
        """
        :return: Pressures measurements from front patch
        """
        _pressure_front = np.zeros(shape=(48, self.num_measurements))
        row_count = 0
        for row in self.sheet.iter_rows(
                min_row=4,
                min_col=1,
                max_row=min(self.num_sensors, 48) + 3,
                max_col=self.num_measurements,
                values_only=True):
            column_count = 0
            for value in row:
                value = float(value)
                if value > self.max_pressure:
                    self.max_pressure = value
                elif value < 0:
                    value = 0
                _pressure_front[row_count, column_count] = value
                column_count += 1
            row_count += 1
        return _pressure_front

    def read_pressure_back_top(self):
        """
        :return: Pressures measurements from back top patch
        """
        _pressure_back_top = np.zeros(shape=(22, self.num_measurements))
        if self.num_sensors > 48:
            # Read pressure values for the top back array
            row_count = 0
            for row in self.sheet.iter_rows(
                    min_row=49 + 3,
                    min_col=1,
                    max_row=min(self.num_sensors, 70) + 3,
                    max_col=self.num_measurements,
                    values_only=True):
                column_count = 0
                for value in row:
                    value = float(value)
                    if value > self.max_pressure:
                        self.max_pressure = value
                    elif value < 0:
                        value = 0
                    _pressure_back_top[row_count, column_count] = value

                    column_count += 1
                row_count += 1
        return _pressure_back_top

    def read_pressure_back_bottom(self):
        _pressure_back_bottom = np.zeros(shape=(10, self.num_measurements))
        if self.num_sensors > 70:
            row_count = 0
            for row in self.sheet.iter_rows(
                    min_row=71 + 3,
                    min_col=1,
                    max_row=self.num_sensors + 3,
                    max_col=self.num_measurements,
                    values_only=True):
                column_count = 0
                for value in row:
                    value = float(value)
                    if value > self.max_pressure:
                        self.max_pressure = value
                    elif value < 0:
                        value = 0
                    _pressure_back_bottom[row_count, column_count] = value
                    column_count += 1
                row_count += 1

        return _pressure_back_bottom
