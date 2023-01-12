import math
import os
import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from recording.plotter import Plot


class LabelDataAnalysis:
    # Process all samples for one given label

    def __init__(self, data_folder, label):
        """
        Initialize the class with the data folder and label
        :param data_folder: path to the folder containing the data files
        :param label: label for the data (e.g correct_orthosis, incorrect_orthosis, no_orthosis)
        """
        self.data_folder = data_folder
        self.dir = os.listdir(self.data_folder)
        self.label = label

    def mean_variance_at_degrees(self, at_arm_degrees, linearized=True):
        """
        Compute the mean and variance of pressure and rotation sensor values at given arm degrees
        :param at_arm_degrees: list of arm degrees to calculate the mean and variance at
        :param linearized: flag to indicate if the data should be linearized or not
        :return: tuple of mean and variance of pressure and rotation sensor values
        """
        p_sensor_values = []
        rot_sensor_values = []
        for file in self.dir:
            if (linearized and "LIN" not in file) or (not linearized and "RAW" not in file):
                continue

            path = self.data_folder + "/" + file
            sample = SampleDataAnalysis(self.label, file_path=path)

            p_sensors, rot_sensors, time_stamp = sample.sensor_values_at_arm_degrees(at_arm_degrees)
            p_sensor_values.append(p_sensors)
            rot_sensor_values.append(rot_sensors)

        # Computes the MEAN for pressure values
        p_sensor_avg = np.zeros(len(p_sensor_values[0]))
        for sensor_values in p_sensor_values:
            p_sensor_avg += sensor_values
        p_sensor_avg = p_sensor_avg / len(p_sensor_values)

        # Computes the MEAN for rotation angle values
        rot_sensor_avg = np.zeros(len(rot_sensor_values[0]))
        for sensor_values in rot_sensor_values:
            rot_sensor_avg += sensor_values
        rot_sensor_avg = rot_sensor_avg / len(rot_sensor_values)

        # Computes the VARIANCE for pressure values
        p_sensor_var = np.zeros(len(p_sensor_values[0]))
        for sensor_values in p_sensor_values:
            p_sensor_var += (p_sensor_avg - sensor_values) ** 2
        p_sensor_var = p_sensor_var / len(p_sensor_values)

        # Computes the VARIANCE for rotation angle values
        rot_sensor_var = np.zeros(len(rot_sensor_values[0]))
        for sensor_values in rot_sensor_values:
            rot_sensor_var += (rot_sensor_avg - sensor_values) ** 2
        rot_sensor_var = rot_sensor_var / len(rot_sensor_values)

        return p_sensor_avg, p_sensor_var, rot_sensor_avg, rot_sensor_var


class SampleDataAnalysis:
    # Hardcoded index ranges of each of the three sensors arrays on the arm
    front_range = (0, 48)
    top_range = (48, 70)
    bot_range = (70, 80)

    def __init__(self, label, file_path=None, p_sensors=None, rot_sensors=None, sample_time_stamps=None):
        """
        Either initialize from csv file, or from prebuilt arrays

        :param label: label for this sample
        :param file_path: path to this sample's data
        :param p_sensors: The pressure sensor values for this sample
        :param rot_sensors: The rotation sensor values for this sample
        :param sample_time_stamps: The time stamps for this sample
        """
        self.label = label
        if file_path is not None:
            self.file_path = file_path
            self.sheet = openpyxl.load_workbook(self.file_path).active
            self.p_sensors, self.rot_sensors, self.sample_time_stamps = self.load_sample()

        if p_sensors is not None and sample_time_stamps is not None and rot_sensors is not None:
            self.p_sensors = p_sensors
            self.rot_sensors = rot_sensors
            self.sample_time_stamps = sample_time_stamps

        self.time_step = self.sample_time_stamps[1] - self.sample_time_stamps[0]
        self.num_p_sensors = len(self.p_sensors[0])
        self.num_rot_sensors = len(self.rot_sensors[0])
        self.num_time_steps = len(self.sample_time_stamps)

    def load_sample(self):
        """
        Structure of the Excel sheet:
            The first row contains the column headers
            The second row contains the sample time stamps
            The last two rows contain the rotation sensor values
            The remaining rows contain the pressure sensor values
        """

        num_sensors = self.sheet.max_row
        num_columns = self.sheet.max_column

        # The time stamps for each sample
        sample_time_stamps = np.zeros(num_columns)
        # The pressure sensor values for each sample
        p_sensors = np.zeros((num_columns, num_sensors - 6))
        # The rotation sensor values for each sample
        rot_sensors = np.zeros((num_sensors, 2))

        for i in range(num_columns):
            sample_time_stamps[i] = self.sheet.cell(2, i + 1).value
            rot_sensors[i][0] = self.sheet.cell(num_sensors - 1, i + 1).value
            rot_sensors[i][1] = self.sheet.cell(num_sensors, i + 1).value
            for j in range(0, num_sensors - 6):
                p_sensors[i][j] = self.sheet.cell(j + 4, i + 1).value

        return p_sensors, rot_sensors, sample_time_stamps

    def sensor_values_at_arm_degrees(self, at_arm_degrees, range_=(0, 80)):
        """
        :param at_arm_degrees: The arm degree for which the sensor values are requested
        :param range_: The range of the pressure sensor values to be returned
        :return: pressure and rotation sensor values at a given arm degree
        """

        min_difference = abs(at_arm_degrees - self.rot_sensors[0][0])
        index = 0

        # Compare absolute difference between the given arm degree and the rotation sensor value
        # at each time step with the current minimum difference
        for i in range(1, self.num_time_steps):
            difference = abs(at_arm_degrees - self.rot_sensors[i][0])
            if difference < min_difference:
                index = i
                min_difference = difference
            if difference < 10e-5:
                break

        return self.p_sensors[index][range_[0]:range_[1]], self.rot_sensors[index], self.sample_time_stamps[index]

    def sensor_wise_avg(self, avg_range=None):
        """
        :param avg_range: a range of time steps for which the average is to be calculated.
        (If no range is provided, it defaults to averaging over all the time steps)
        :return: the average of pressure and rotation sensor values over a given range
        """

        if avg_range is None:
            avg_range = range(self.num_time_steps)
        else:
            avg_range = range(max(0, avg_range[0]), min(self.num_time_steps, avg_range[-1] + 1))

        p_sensors_avg = np.zeros(self.num_p_sensors)
        rot_sensors_avg = np.zeros(2)
        for i in avg_range:
            p_sensors_avg += self.p_sensors[i]
            rot_sensors_avg += self.rot_sensors[i]

        return p_sensors_avg / len(avg_range), rot_sensors_avg / len(avg_range)

    def timewise_avg(self):
        """
        :return: the average of pressure and rotation sensor values for each time stamp
        """
        p_time_stamp_avg = np.zeros(self.num_time_steps)
        rot_time_stamp_avg = np.zeros(self.num_time_steps)

        # go over all the time steps, for each calculate the average of pressure and rotation sensor values
        for i in range(self.num_time_steps):
            p_time_stamp_avg[i] = sum(self.p_sensors[i]) / self.num_p_sensors
            rot_time_stamp_avg[i] = sum(self.rot_sensors[i]) / self.num_rot_sensors

        return p_time_stamp_avg, rot_time_stamp_avg

    def total_avg(self):
        """
        :return: the total average of pressure and rotation sensor values
        """
        p_avg, rot_avg = self.sensor_wise_avg(range(0, self.num_time_steps))
        return sum(p_avg) / len(p_avg), sum(rot_avg) / len(rot_avg)

    def avg_at_arm_degrees(self, at_arm_degrees, range_=(0, 80)):
        """
        :param at_arm_degrees: The arm degree for which the sensor values are requested
        :param range_: The range of the pressure sensor values to be returned
        :return: the average of pressure and rotation sensor values at a given arm degree
        """
        p_values, rot_values, time_stamp = self.sensor_values_at_arm_degrees(at_arm_degrees, range_=range_)

        return sum(p_values) / len(p_values), sum(rot_values) / len(rot_values)

    def get_p_value(self, column, row):
        """
        :return: the pressure sensor value at a given column and row
        """
        return self.p_sensors[column][row]

    def get_rot_values(self, column, row):
        """
        :return: the rotation sensor value at a given column and row
        """
        return self.rot_sensors[column][row]

    def get_p_values_at_time(self, column):
        """
        :return: all the pressure sensor values
        """
        return self.p_sensors[column]

    def get_rot_values_at_time(self, column):
        """
        :return: the rotation sensor values for time stamp from given column
        """
        return self.rot_sensors[column]

    def get_time_at(self, column):
        """
        :return: time stamp from given column
        """
        return self.sample_time_stamps[column]

    def get_sensor_values(self, sensor):
        return self.p_sensors[:, sensor]

    def shrink_sample_time_domain(self, k_shrink):
        """
        Makes sample smaller in time domain
        for each time stamp computes the average of the timestamps that fall into the new time stamp range
        :param k_shrink: number of new time stamps
        """

        shrink_ratio = math.ceil(self.num_time_steps / k_shrink)

        new_p_sensors = np.zeros((k_shrink, self.num_p_sensors))
        new_rot_sensors = np.zeros((k_shrink, self.num_rot_sensors))
        new_sample_time_stamps = np.zeros(k_shrink)

        for i in range(k_shrink):
            new_p_sensors[i], new_rot_sensors[i] = self.sensor_wise_avg(range(i * shrink_ratio, (i + 1) * shrink_ratio))
            new_sample_time_stamps[i] = self.sample_time_stamps[i * shrink_ratio]

        return SampleDataAnalysis(
            p_sensors=new_p_sensors, rot_sensors=new_rot_sensors,
            sample_time_stamps=new_sample_time_stamps, label=self.label
        )

    def extrema_pressure_time_stamp(self, in_range=None):
        """

        :param in_range: a range of time steps
        :return: the pressure sensor values at the time step
        with the minimum pressure and at the time step with the maximum pressure
        """

        if in_range is None:
            in_range = range(self.num_time_steps)

        max_i = in_range[0]
        max_press = sum(self.p_sensors[in_range[0]])
        min_i = in_range[0]
        min_press = sum(self.p_sensors[in_range[0]])

        for i in in_range:
            p_sum = sum(self.p_sensors[i])
            if p_sum < min_press:
                min_i, min_press = i, p_sum
            if p_sum > max_press:
                max_i, max_press = i, p_sum

        return self.p_sensors[min_i], self.p_sensors[max_i]

    def extrema_pressure_sensor(self):
        """
        :return: the pressure sensor values for the sensor with the minimum pressure
        and for the sensor with the maximum pressure.
        """

        min_i, max_i = 0, 0
        min_press = sum(self.p_sensors[:, 0])
        max_press = sum(self.p_sensors[:, 0])

        for i in range(1, self.num_p_sensors):
            p_sum = sum(self.p_sensors[:, i])
            if p_sum < min_press:
                min_i, min_press = i, p_sum
            if p_sum > max_press:
                max_i, max_press = i, p_sum

        return self.p_sensors[:, min_i], self.p_sensors[:, max_i]

    def copy(self):
        return SampleDataAnalysis(
            self.label, p_sensors=np.copy(self.p_sensors), rot_sensors=np.copy(self.rot_sensors),
            sample_time_stamps=np.copy(self.sample_time_stamps)
        )


def plot_pressure_data(sensor_array):
    plot = Plot(10)
    plt.pause(0.05)
    plot.update_pressure_plot(sensor_array[1:49], sensor_array[49:71], sensor_array[71:])
    plt.pause(20)


def plot_pressure_degree_relation(path, label, max_deg, range_=(0, 80)):
    data = LabelDataAnalysis(path + "/" + label, label)
    scale = 0.1

    deg_range, pressures, stds = [], [], []
    for d in range(max_deg + 1):
        deg_range.append(d)
        p_sensor_avg, p_sensor_var, rot_sensor_avg, rot_sensor_var = data.mean_variance_at_degrees(d, linearized=False)
        pressures.append(np.mean(p_sensor_avg[range_[0]:range_[1]]))
        stds.append(np.std(p_sensor_avg[range_[0]:range_[1]]) * scale)

    plt.errorbar(deg_range, pressures, yerr=stds, fmt="o", color="black", ecolor="palevioletred", elinewidth=2)
    plt.xlabel("degrees of rotation")
    plt.xlabel("average pressure")
    plt.title(label)
    plt.show()


if __name__ == "__main__":
    DATA_PATH = "C:/Users/majag/Desktop/marble/NewData"

    for folder in os.listdir(DATA_PATH):
        print(int(folder.split("_")[-1]))
        plot_pressure_degree_relation(
            DATA_PATH,
            label=folder,
            max_deg=int(folder.split("_")[-1]),
            range_=SampleDataAnalysis.front_range
        )
