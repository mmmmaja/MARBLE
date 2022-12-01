import os
import sys

import numpy as np
import openpyxl
import xlsxwriter

path = "C:/University/Marble/Data/"

data_dir = os.listdir(path)



## Code to add angle values also to the raw files



for label in data_dir:

    label_path = path + label
    dir = os.listdir(label_path)
    n = int(len(dir)/2)


    print("DIR: ",label_path)
    print("folder size:", n)
    for k in range(n):
        r_path = "None"
        try:
            f_name = f"{k+1}_RAW.xlsx"
            r_path = label_path + "/" + f_name
            sample = np.zeros((83,20))
            r_c= openpyxl.load_workbook(r_path)
            raw = r_c.active


            for i in range(20):
                for j in range(80):
                    sample[j+1][i] = raw.cell(row = j+4, column= i+1).value
                sample[0][i] = raw.cell(row = 2,column=i+1).value



            l_path = label_path + f"/{k + 1}_LIN.xlsx"
            l_c = openpyxl.load_workbook(l_path)
            lin = l_c.active





            for i in range(20):
                sample[81][i] = lin.cell(row = 85,column=i+1).value
                sample[82][i] = lin.cell(row = 86,column=i+1).value

            r_c.close()
            l_c.close()

            workbook = xlsxwriter.Workbook(r_path)
            worksheet = workbook.add_worksheet()
            worksheet.write(0, 0, str(k+1))

            for col_num, data in enumerate(sample[0]):
                worksheet.write(1, col_num, data)
                worksheet.write(84,col_num, sample[81][col_num])
                worksheet.write(85,col_num, sample[82][col_num])

            for i in range(80):
                index = i + 1
                for col_num, data in enumerate(sample[index]):
                    worksheet.write(index + 2, col_num, data)


            workbook.close()
            print(r_path)
        except:

            print("[ WARNING ] ",r_path, "DOES NOT EXIST")


    print("\n\n")



