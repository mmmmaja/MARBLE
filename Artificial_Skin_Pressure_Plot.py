import numpy as np
import matplotlib.pyplot as plt
import matplotlib.tri as tri
from matplotlib.tri import triangulation, LinearTriInterpolator
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter,
                               AutoMinorLocator)
from mpl_toolkits.mplot3d import Axes3D
#from mpl_toolkits.mplot3d.art3d import Poly3DCollection
from pylab import get_cmap
from scipy.interpolate import LinearNDInterpolator
import cv2
import openpyxl
from pathlib import Path
import math

forceLimit = 2 #specify the force limit in Newtons
plot = 1        # '1'= 2D contour plot, '2'= 2D multilayer plt, '3'=3D plot
timeStep=500  #time step in ms with which the plot is updated. Must be multiple of the recording time step (50ms)
plotDelay=0.5 #delays updating the 2D contour plot [ms]
plot_patch02 =0 # set to '1' if you want to plot array patch02 (ventral top)
plot_patch03 =0 # set to '1' if you want to plot array patch03 (ventral bottom)
xlsx_file = Path('data', 'incorrect_orthosis_up_1cm_45', '1_LIN.xlsx') #path of the xlsx file to read

#node (sensor) coordinates ventral (plus pressure dummy data)
nds_frontX = [0.000, 25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000,  25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000,  25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000,  25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000,  25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000,  25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933]
nds_frontY = [0.000, 0.000,  0.000,  0.000,  7.500,  7.500,  7.500,  7.500,  15.000, 15.000, 15.000, 15.000, 22.500, 22.500, 22.500, 22.500, 30.000, 30.000, 30.000, 30.000, 37.500, 37.500, 37.500, 37.500, 45.000, 45.000, 45.000, 45.000, 52.500, 52.500, 52.500, 52.500, 60.000, 60.000, 60.000, 60.000, 67.500, 67.500, 67.500, 67.500, 75.000, 75.000, 75.000, 75.000, 82.500, 82.500, 82.500, 82.500 ]
nds_frontZ=  [5.50, 2.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.200,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000 ]
pressures = [10.000, 15.000,  0.000,  0.000,  5.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000,  0.000 ]

#node (sensor) coordinates dorsal top
nds_backTopX= [0.000, 25.981, 51.962, 77.943, 103.924, 129.905, 12.990, 38.971, 64.952, 90.933, 116.915, 0.000,  25.981, 51.962, 77.943, 103.924, 129.905, 12.990, 38.971, 64.952, 90.933, 116.915]
nds_backTopY= [0.000, 0.000,  0.000,  0.000,  0.000,   0.000,   7.500,  7.500,  7.500,  7.500,  7.500,   15.000, 15.000, 15.000, 15.000, 15.000,  15.000,  22.500, 22.500, 22.500, 22.500, 22.500]

#node (sensor) coordinates dorsal bottom
nds_backBottomX=[0.000, 25.981, -12.990, 12.990, 38.971, 0.000, 25.981,  -12.990, 12.990, 38.971]
nds_backBottomY=[0.000, 0.000,   7.500,  7.500,  7.500,  15.000, 15.000,  22.500, 22.500, 22.500]
###########read and parse sensor data from xlsx file###########################
wb_obj = openpyxl.load_workbook(xlsx_file) 
# Read the active sheet:
sheet = wb_obj.active
num_sensors=sheet.max_row-6
num_measurements=sheet.max_column
#pressure_front=np.zeros(shape=(num_sensors,sheet.max_column))
pressure_front=np.zeros(shape=(48,num_measurements)) #array with zeros for pressure values of the front sensor array
pressure_backTop=np.zeros(shape=(22,num_measurements))
pressure_backBottom=np.zeros(shape=(10,num_measurements))
time=np.zeros(num_measurements)
angle_arm=np.zeros(num_measurements)
angle_orthosis=np.zeros(num_measurements)
print("Number of Sensors " + str(num_sensors))
print("Number of Measurements: " + str(num_measurements))

#read time values
for i in range(1,num_measurements):
    time[i] = sheet.cell(column=i, row=2).value #time in secs
timeStepRec=(time[1]-time[0])*1000 #get recording time step in [ms]
print("Recording time step: " + str(timeStepRec))

#read arm angle
for i in range(1,num_measurements):
    angle_orthosis[i] = sheet.cell(column=i, row=num_sensors+6).value #time in secs

#Arm angles below zero or above 1000 are 0 (happens if arm is outside the calibration range)
angle_orthosis[angle_orthosis>1000]=0
angle_orthosis[angle_orthosis<0]=0
#print("Orthosis angle")
#print(angle_orthosis)

#read orthosis angle
for i in range(1,num_measurements):
    angle_arm[i] = sheet.cell(column=i, row=num_sensors+5).value #time in secs

#Arm angles above 1000 are 0 (happens if arm is outside the calibration range)
angle_arm[ angle_arm>1000]=0
#print("Arm angle")
#print(angle_arm)

#print("Recording time step in [ms]")
#print(timeStepRec)

#read pressure values
row_count=0
column_count=0
if num_sensors<=48:
    for row in sheet.iter_rows(min_row=4, min_col=1, max_row=num_sensors+3, max_col=num_measurements, values_only=True):
        column_count=0
        for value in row:
            pressure_front[row_count,column_count]=value
            column_count=column_count+1
        row_count=row_count+1

    #For now replace negative pressure values by zero
    pressure_front[pressure_front<0]=0

elif num_sensors>48 and num_sensors<=70:
    #read pressure values for the front array
    for row in sheet.iter_rows(min_row=4, min_col=1, max_row=48+3, max_col=num_measurements, values_only=True):
        column_count=0
        for value in row:
            pressure_front[row_count,column_count]=value
            column_count=column_count+1
        row_count=row_count+1
    
    #read pressure values for the top back array
    row_count=0
    for row in sheet.iter_rows(min_row=49+3, min_col=1, max_row=num_sensors+3, max_col=num_measurements, values_only=True):
        column_count=0
        for value in row:
            pressure_backTop[row_count,column_count]=value
            column_count=column_count+1
        row_count=row_count+1


    #For now replace negative pressure values by zero
    pressure_front[pressure_front<0]=0
    pressure_backTop[pressure_backTop<0]=0

elif num_sensors>70:
    #read pressure values for the front array
    for row in sheet.iter_rows(min_row=4, min_col=1, max_row=48+3, max_col=num_measurements, values_only=True):
        column_count=0
        for value in row:
            pressure_front[row_count,column_count]=value
            column_count=column_count+1
        row_count=row_count+1
    
    #read pressure values for the top back array
    row_count=0
    for row in sheet.iter_rows(min_row=49+3, min_col=1, max_row=70+3, max_col=num_measurements, values_only=True):
        column_count=0
        for value in row:
            pressure_backTop[row_count,column_count]=value
            column_count=column_count+1
        row_count=row_count+1

    #read pressure values for the back bottom array
    row_count=0
    for row in sheet.iter_rows(min_row=71+3, min_col=1, max_row=num_sensors+3, max_col=num_measurements, values_only=True):
        column_count=0
        for value in row:
            pressure_backBottom[row_count,column_count]=value
            column_count=column_count+1
        row_count=row_count+1


    #For now replace negative pressure values by zero
    pressure_front[pressure_front<0]=0
    pressure_backTop[pressure_backTop<0]=0
    pressure_backBottom[pressure_backBottom<0]=0
    #print("Pressure Front")
    #print(pressure_front)
    #print("Pressure top back")
    #print(pressure_backTop)
############ DEFINE  triangles_front #################################
#always counter clockwise 
triangles_front= [
    [4, 1, 0],
    [5, 2, 1],
    [6, 3, 2],
    [8, 4, 0],
    [9, 1, 4],
    [9, 5, 1],
    [10, 2, 5],
    [10, 6, 2],
    [11, 3, 6],
    [11, 7, 3],
    [12, 4, 8],
    [12, 9, 4],
    [13, 5, 9],
    [13, 10, 5],
    [14, 6, 10],
    [14, 11, 6],
    [15, 7, 11],
    [16, 12, 8],
    [17, 9, 12],
    [17, 13, 9],
    [18, 10, 13],
    [18, 14, 10],
    [19, 11, 14],
    [19, 15, 11],
    [20, 12, 16],
    [20, 17, 12],
    [21, 13, 17],
    [21, 18, 13],
    [22, 14, 18],
    [22, 19, 14],
    [23, 15, 19],
    [24, 20, 16],
    [25, 17, 20],
    [25, 21, 17],
    [26, 18, 21],
    [26, 22, 18],
    [27, 19, 22],
    [27, 23, 19],
    [28, 20, 24],
    [28, 25, 20],
    [29, 21, 25],
    [29, 26, 21],
    [30, 22, 26],
    [30, 27, 22],
    [31, 23, 27],
    [32, 28, 24],
    [33, 25, 28],
    [33, 29, 25],
    [34, 26, 29],
    [34, 30, 26],
    [35, 27, 30],
    [35, 31, 27],
    [36, 28, 32],
    [36, 33, 28],
    [37, 29, 33],
    [37, 34, 29],
    [38, 30, 34],
    [38, 35, 30],
    [39, 31, 35],
    [40, 36, 32],
    [41, 33, 36],
    [41, 37, 33],
    [42, 34, 37],
    [42, 38, 34],
    [43, 35, 38],
    [43, 39, 35],
    [44, 36, 40],
    [44, 41, 36],
    [45, 37, 41],
    [45, 42, 37],
    [46, 38, 42],
    [46, 43, 38],
    [47, 39, 43],
    [45, 41, 44],
    [46, 42, 45],
    [47, 43, 46]
]
#triangles of the dorsdal top array
triangles_backTop= [
    [6, 1, 0],
    [7, 2, 1],
    [8, 3, 2],
    [9, 4, 3],
    [10, 5, 4],
    [11, 6, 0],
    [12, 1, 6],
    [12, 7, 1],
    [13, 2, 7],
    [13, 8, 2],
    [14, 3, 8],
    [14, 9, 3],
    [15, 4, 9],
    [15, 10, 4],
    [16, 5, 10],
    [17, 6, 11],
    [17, 12, 6],
    [18, 7, 12],
    [18, 13, 7],
    [19, 8, 13],
    [19, 14, 8],
    [20, 9, 14],
    [20, 15, 9],
    [21, 10, 15],
    [21, 16, 10],
    [17, 18, 12],
    [18, 19, 13],
    [19, 20, 14],
    [20, 21, 15]
    ]
#triangles of the dorsal bottom array
triangles_backBottom= [
    [3, 1, 0],
    [5, 0, 2],
    [5, 3, 0],
    [6, 1, 3],
    [6, 4, 1],
    [7, 5, 2],
    [8, 3, 5],
    [8, 6, 3],
    [9, 4, 6],
    [7, 8, 5],
    [8, 9, 6]
    ]

########################### Int for 2D Plot ##############################
#triangulate arrays
tri_front = tri.Triangulation(nds_frontX, nds_frontY,triangles_front) #triangulation of the ventral array
tri_backTop = tri.Triangulation(nds_backTopX, nds_backTopY,triangles_backTop) #triangulation of the ventral array
tri_backBottom = tri.Triangulation(nds_backBottomX, nds_backBottomY,triangles_backBottom) #triangulation of the ventral array

#convert arrays to numpy arrays
#Front
nds_frontX=np.array(nds_frontX)
nds_frontY=np.array(nds_frontY)
nds_frontZ=np.array(nds_frontZ)
pressures=np.array(pressures)
#Back top
nds_backTopX=np.array(nds_backTopX)
nds_backTopY=np.array(nds_backTopY)
#Back top
#nds_backBottomX=np.array(nds_backBottomX)
#nds_backBottomY=np.array(nds_backBottomY)

################################ Init fOR 3D PlOT ####################################
#define color pallet for 3D plot
colorTable =[[0,0,255], [0,128,255], [0,255,255],[0,255,128],[0,255,0],[128,255,0],[255,255,0],[255,128,0],[255,0,0]]

##for 3d plot init arrays
presScaled=np.zeros(len(pressures))
colorIdx =np.zeros(len(pressures),dtype=np.uint8)
nodeColors =np.zeros(shape=(len(pressures),3),dtype=np.uint8)

test =np.zeros((2,3), dtype=np.uint8)
array = np.arange(9).reshape(3,3)
test2 =[[0,0,0], [0,0,0], [0,0,0]]

# create black background image
#result = np.zeros((len(triangles_front),int(max(nds_frontX)),int(max(nds_frontX)),3), dtype=np.uint8) #triangles_front,y,x,rgb
result = np.zeros((len(triangles_front),105,105,3), dtype=np.uint8) #triangles_front,y,x,rgb
triArr_xy = np.zeros((len(triangles_front),3,3), dtype=np.uint8)
triArr_z = np.zeros((len(triangles_front),1,3), dtype=np.uint8)

coords_x=np.zeros((len(triangles_front),105,105), dtype=np.uint8)
coords_y=np.zeros((len(triangles_front),105,105), dtype=np.uint8)
coords_z=np.zeros((len(triangles_front),105,105), dtype=np.uint8)
#coords_x=np.zeros((len(triangles_front),int(max(nds_frontX)),int(max(nds_frontX))), dtype=np.uint8)
#coords_y=np.zeros((len(triangles_front),int(max(nds_frontX)),int(max(nds_frontX))), dtype=np.uint8)
#coords_z=np.zeros((len(triangles_front),int(max(nds_frontX)),int(max(nds_frontX))), dtype=np.uint8)


#Create a color scale and assign colors to the node pressures
for i in range(len(pressures)):
    # Specify colors
    if max(pressures) > forceLimit:
        presScaled[i]= float(pressures[i]/max(pressures))
        colorIdx[i] = int(round(presScaled[i]*8))
        #print(colorIdx[i])
        #print(colorTable[colorIdx[i]])
        nodeColors[i]=colorTable[colorIdx[i]]
    else:
        presScaled[i]= float(pressures[i]/forceLimit)
        colorIdx[i] = int(round(presScaled[i]*8))
        nodeColors[i]=colorTable[colorIdx[i]]


# Make array of vertices
# ax bx cx
# ay by cy
#  1  1  1
for t in range(len(triangles_front)):
    triArr_xy[t] = np.asarray([nds_frontX[triangles_front[t][0]], nds_frontX[triangles_front[t][1]], nds_frontX[triangles_front[t][2]], nds_frontY[triangles_front[t][0]], nds_frontY[triangles_front[t][1]], nds_frontY[triangles_front[t][2]], 1,1,1]).reshape((3,3))
    triArr_z[t] = np.asarray([nds_frontZ[triangles_front[t][0]], nds_frontZ[triangles_front[t][1]], nds_frontZ[triangles_front[t][2]]])
#print("tri_array")
#print(triArr_z)
x1=0
y1=0
x1_prev=0
y1_prev=0

for t in range(len(triangles_front)):

    #Debugging: for getting array range
    if x1 > x1_prev:
        x1_prev=x1
    if y1 > y1_prev:
        y1_prev=y1

    x1=0
    y1=0

    # Get bounding box of the triangle
    xleft = min(nds_frontX[triangles_front[t][0]], nds_frontX[triangles_front[t][1]], nds_frontX[triangles_front[t][2]])
    xright = max(nds_frontX[triangles_front[t][0]], nds_frontX[triangles_front[t][1]], nds_frontX[triangles_front[t][2]])
    ytop = min(nds_frontY[triangles_front[t][0]], nds_frontY[triangles_front[t][1]], nds_frontY[triangles_front[t][2]])
    ybottom = max(nds_frontY[triangles_front[t][0]], nds_frontY[triangles_front[t][1]], nds_frontY[triangles_front[t][2]])

    #xleft=int(xleft)
    #xright=int(xright)
    #ytop=int(ytop)
    #ybottom=int(ybottom)

    # loop over each pixel, compute barycentric coordinates and interpolate vertex colors
    for y in np.arange(ytop, ybottom,1):
        
        for x in np.arange(xleft, xright,1):
            # Store the current point as a matrix
            p = np.array([[x], [y], [1]])
            #print("----------P---------------")
            #print(p)
            #print("-------------------------")
            #print("----------Tri---------------")
            #print(t)
            #print(triArr_xy[t,0])
            #print(triArr_xy[t,1])
            #print("-------------------------")
            
            # Solve for least squares solution to get barycentric coordinates
            (alpha, beta, gamma) = np.linalg.lstsq(triArr_xy[t], p, rcond=-1)[0]

            # The point is inside the triangle if all the following conditions are met; otherwise outside the triangle
            if alpha > 0 and beta > 0 and gamma > 0:
                # do barycentric interpolation on colors

                color=nodeColors[triangles_front[t][0]]*alpha + nodeColors[triangles_front[t][1]]*beta + nodeColors[triangles_front[t][2]]*gamma

                #coords_x[t][y,x]= triArr_xy[t,0,0] + ( triArr_xy[t,0,1] - triArr_xy[t,0,0]) * alpha  + (triArr_xy[t,0,2] - triArr_xy[t,0,0]) * beta
                #coords_y[t][y,x]= triArr_xy[t,1,0] + ( triArr_xy[t,1,1] - triArr_xy[t,1,0]) * alpha + (triArr_xy[t,1,2] - triArr_xy[t,1,0]) * beta
                coords_z[t][y1,x1]= triArr_z[t,0,0] + ( triArr_z[t,0,1] - triArr_z[t,0,0]) *beta + (triArr_z[t,0,2] - triArr_z[t,0,0]) *gamma

                result[t][y1,x1] = color
                coords_x[t][y1,x1]=p[0]
                coords_y[t][y1,x1]=p[1] #p.reshape((3,))     # p.shape  = (3,) #store x coord
                x1=x1+1
                y1=y1+1

################################################ 3D Plot ##########################################
if plot==3:
    #print("x1 prev")
    #print(x1_prev)
    #print(y1_prev)

    test = np.array(result[0,:,:])
    #test1=test.reshape(8100,3)
    test1=test.reshape(11025,3)
    print(test1.shape)
    C=test1
    print("____")
    print(C)
    print(C.shape)
    print("____")

    #print("x-coordinates")
    #print(triArr_xy[0,0,0])
    #print(triArr_xy[0,0,1])
    #print(triArr_xy[0,0,2])
    #print("y-coordinates")
    #print(triArr_xy[0,1,0])
    #print(triArr_xy[0,1,1])
    #print(triArr_xy[0,1,2])
    #print("z-coordinates")
    #print(triArr_z[0,0,0])
    #print(triArr_z[0,0,1])
    #print(triArr_z[0,0,2])
    #print("____")

    colorN0=nodeColors[triangles_front[0][0]]
    colorN1=nodeColors[triangles_front[0][1]]
    colorN2=nodeColors[triangles_front[0][2]]

    print(coords_x.shape)
    print(coords_y.shape)
    print(coords_z.shape)
    #fig = plt.figure()
    ax = plt.subplot(111, projection = '3d')
    cmap = plt.get_cmap('jet')
    #ax.scatter(coords_x[0,:],coords_y[0,:],0, c = C/255)
    ax.scatter(coords_x[0,:],coords_y[0,:],coords_z[0,:], c = C/255)
    #ax.scatter(triArr_xy[0,0,0],triArr_xy[0,1,0],1,s=150, c=colorN0/255, alpha=0.9) #one corner point of the triangle
    #ax.scatter(triArr_xy[0,0,1],triArr_xy[0,1,1],1,s=150, c=colorN1/255, alpha=0.9) #one corner point of the triangle
    #ax.scatter(triArr_xy[0,0,2],triArr_xy[0,1,2],1,s=150, c=colorN2/255, alpha=0.9) #one corner point of the triangle
    pa=ax.scatter(triArr_xy[0,0,0],triArr_xy[0,1,0],triArr_z[0,0,0],s=100, c=colorN0/255, alpha=0.9) #one corner point of the triangle
    ax.scatter(triArr_xy[0,0,1],triArr_xy[0,1,1],triArr_z[0,0,1],s=100, c=colorN1/255, alpha=0.9) #one corner point of the triangle
    ax.scatter(triArr_xy[0,0,2],triArr_xy[0,1,2],triArr_z[0,0,2],s=100, c=colorN2/255, alpha=0.9) #one corner point of the triangle
    ax.plot_trisurf(tri_front, nds_frontZ,  edgecolor = 'grey', shade=False, linewidth=0.5,alpha=0.01)
    ax.set_xlabel('X-axis', fontweight ='bold')
    ax.set_ylabel('Y-axis', fontweight ='bold')
    ax.set_zlabel('Z-axis', fontweight ='bold')
    ax.set_zlim3d(0, 10)
    plt.colorbar(pa, ax = ax, shrink = 0.5, aspect = 5)
    pa.set_clim(vmin=0, vmax=forceLimit)
    plt.show()

    # save results
    cv2.imwrite('barycentric_triange.png', result)

################################################################################################
#3D mesh with face colors being the meanof the triangle vertex colors
#colors = np.mean(pressures[triangles_front], axis=1)
#print(colors.shape) 

## Plotting
#fig = plt.figure()
#ax = plt.axes(projection='3d')
#cmap = plt.get_cmap('jet')
#collec = ax.plot_trisurf(tri_front, nds_frontZ, cmap=cmap, edgecolor = 'grey', shade=False, linewidth=0.5)
#scatter =ax.scatter(nds_frontX, nds_frontY, nds_frontZ, marker='.', cmap=cmap,s=100, c=pressures, alpha=0.9)
#collec.set_array(colors)
##scatter.set_array(colors)

#collec.autoscale()
#ax.set_zlim3d(0, 3)
### Adding labels
#ax.set_xlabel('X-axis', fontweight ='bold')
#ax.set_ylabel('Y-axis', fontweight ='bold')
#ax.set_zlabel('Z-axis', fontweight ='bold')
#fig.colorbar(collec, ax = ax, shrink = 0.5, aspect = 5)
#plt.show()

################################################ 2D (contour) Plot ##########################################
if plot==1:
    #define colorbar levels and step size
    max_presFront=max(max(x) for x in pressure_front) #max pressure on ventral array
    max_presBackTop=max(max(x) for x in pressure_backTop) #max pressure on dorsal top array
    max_presBackBottom=max(max(x) for x in pressure_backBottom) #max pressure on dorsal top array
    max_presTotal=max(max_presFront,max_presBackTop,max_presBackBottom) #overall max pressure
    print(max_presTotal)
    print("Max recorded force:")
    print(max_presTotal)
    print("Max recorded force on ventral array:")
    print(max_presFront)
    print("Max recorded force on dorsal top array:")
    print(max_presBackTop)
    print("Max recorded force on dorsal bottom array:")
    print(max_presBackBottom)

    #calculate colorbar step size
    if max_presTotal < forceLimit:
        max_presTotal=forceLimit
        steps= 2 * max_presTotal
    else:
        max_presTotal=int(math.ceil(max_presTotal / 10.0)) * 10 #round max force to the next higher decade (for plotting)
        steps= 2*max_presTotal#corresponds to 0.5N steps when havong a force limit of 15N

    #setup plot iteration setp size and labels
    plotStep=int(timeStep/timeStepRec) #calculate the plot update step 
    levels=np.linspace(0,max_presTotal,40) #levels for the colorbar
    level_labels=np.linspace(0,max_presTotal,20) #level labels for the colorbar , endpoint=True
    print("!!!!", max_presTotal)

    plt.ion() #switch on interactive plot

    fig = plt.figure(figsize=(18, 8))
    #2D front triangulation and heatmap of the front array
    ax1 = fig.add_subplot(1,2,1)
    ax1.set_title('Pressure plot - Artificial Skin Array (Ventral). Timestep: %i ms' % timeStep)
    ax1.set_xlabel('Lateral-Medial [mm]',fontsize=12)
    ax1.set_ylabel('Superior-Inferior [mm]',fontsize=12)
    ax1.invert_yaxis()
    ax1.triplot(tri_front, '-k')
    ax1.tick_params(which='major', width=0.75, length=2.5, labelsize=12)
    tcf=ax1.tricontourf(tri_front, pressure_front[:,0],levels,cmap='jet')    

    # Add box showing the time to the plot
    props = dict(boxstyle='round', facecolor='wheat', alpha=0.5) #bounding box of time inidication
    style = dict(size=12, color='black',bbox=props) #"time"-annotation style
    t1=ax1.text(110, 90,'t= 0 ms', **style) #add annotation showing the time

    #add colorbar
    cbar=plt.colorbar(tcf,ax=ax1,ticks=level_labels)
    cbar.ax.set_yticklabels(["%1.1f" % y for y in level_labels], fontsize=12)
    cbar.set_label('Force in [N]', labelpad=-20, y=1.05, rotation=0,fontsize=12) #colorbar label

    if plot_patch02==1:
        #2D  back top triangulation and heatmap of the back top array
        fig2 = plt.figure(figsize=(18, 8))
        ax3 = fig2.add_subplot(1,1,1)
        ax3.set_title('Pressure plot - Artificial Skin Array (Dorsal, Top). Timestep: %i ms' % timeStep)
        ax3.set_xlabel('Lateral-Medial [mm]',fontsize=12)
        ax3.set_ylabel('Superior-Inferior [mm]',fontsize=12)
        ax3.invert_yaxis()
        ax3.triplot(tri_backTop, '-k')
        ax3.tick_params(which='major', width=0.75, length=2.5, labelsize=12)
        ax3.tricontourf(tri_backTop, pressure_backTop[:,0],levels,cmap='jet')    
        t3=ax3.text(130, 25,'t= 0 ms', **style) #add annotation showing the time
        #add colorbar
        cbar=plt.colorbar(tcf,ax=ax3,ticks=level_labels)
        cbar.ax.set_yticklabels(["%1.1f" % y for y in level_labels], fontsize=12)
        cbar.set_label('Force in [N]', labelpad=-20, y=1.05, rotation=0,fontsize=12) #colorbar label
    
    if plot_patch03==1:
        #2D  back top triangulation and heatmap of the back top array
        fig3 = plt.figure(figsize=(18, 8))
        ax4 = fig3.add_subplot(1,1,1)
        ax4.set_title('Pressure plot - Artificial Skin Array (Dorsal, Bottom). Timestep: %i ms' % timeStep)
        ax4.set_xlabel('Lateral-Medial [mm]',fontsize=12)
        ax4.set_ylabel('Superior-Inferior [mm]',fontsize=12)
        ax4.invert_yaxis()
        ax4.triplot(tri_backBottom, '-k')
        ax4.tick_params(which='major', width=0.75, length=2.5, labelsize=12)
        ax4.tricontourf(tri_backBottom, pressure_backBottom[:,0],levels,cmap='jet')    
        t4=ax4.text(40, 25,'t= 0 ms', **style) #add annotation showing the time
        #add colorbar
        cbar=plt.colorbar(tcf,ax=ax4,ticks=level_labels)
        cbar.ax.set_yticklabels(["%1.1f" % y for y in level_labels], fontsize=12)
        cbar.set_label('Force in [N]', labelpad=-20, y=1.05, rotation=0,fontsize=12) #colorbar label
    

    #plot for the arm angle
    ax2 = fig.add_subplot(1,2,2)
    trajectory_arm, =ax2.plot(time,angle_arm,linewidth='3',color='dimgray') #plot arm angle trajectory
    trajectory_orth, =ax2.plot(time,angle_orthosis,linewidth='3',color='limegreen') #plot arm angle trajectory
    point_arm,=ax2.plot(time[0],angle_arm[0],linestyle='none', markerfacecolor='red', marker="D", markeredgecolor="black", markersize=14) #mark arm angle at time zero
    point_orth,=ax2.plot(time[0],angle_arm[0],linestyle='none', markerfacecolor='limegreen', marker="D", markeredgecolor="black", markersize=10) #mark orthosis angle at time zero
    yticks=np.arange(0, 95, step=5)
    ax2.set_yticks(yticks)  # Set label locations.
    ax2.set_xlabel('Time in [secs]',fontsize=12)
    #ax2.set_xticklabels(time, fontsize=12)
    ax2.set_ylabel('Angle [degs]',fontsize=12)
    ax2.set_yticklabels(yticks, fontsize=12)
    # Make a plot with major ticks that are multiples of 5 and minor ticks that
    # are multiples of 1.  Label major ticks with '%d' formatting but don't label
    ax2.xaxis.set_major_locator(MultipleLocator(5))   
    ax2.xaxis.set_major_formatter(FormatStrFormatter('%d'))
    # For the minor ticks, use no labels; default NullFormatter.
    ax2.xaxis.set_minor_locator(MultipleLocator(1))
    ax2.tick_params(which='major', width=0.75, length=2.5, labelsize=12)
    ax2.xaxis.grid(True, which='minor')
    ax2.xaxis.grid(True, which='major')
    ax2.yaxis.grid(True, which='minor')
    ax2.yaxis.grid(True, which='major')
    trajectory_arm.set_label('Arm angle')
    trajectory_orth.set_label('Orthosis angle')
    ax2.legend()
    plt.show()
    plt.pause(plotDelay)

    for k in range(1,num_measurements,plotStep): #num_measurements
        print("Playing Recording")
        ax1.tricontourf(tri_front, pressure_front[:,k],levels,cmap='jet')
        t1.set_text('t= %1.2f secs' % (time[k]))
        #plot array patch02 pressures two if enabled by the user
        if plot_patch02==1:
            ax3.tricontourf(tri_backTop, pressure_backTop[:,k],levels,cmap='jet')
            t3.set_text('t= %1.2f secs' % (time[k]))
        #plot array patch03 pressures two if enabled by the user
        if plot_patch03==1:
            ax4.tricontourf(tri_backBottom, pressure_backBottom[:,k],levels,cmap='jet')
            t4.set_text('t= %1.2f secs' % (time[k]))
        #plot arm and orthosis angles
        point_arm.set_data(time[k],angle_arm[k]) #draw the next point indicating the arm angle
        point_orth.set_data(time[k],angle_orthosis[k]) #draw the next point indicating the orthosis angle
        plt.draw()  
        plt.pause(plotDelay)


##################################################################################################################
#2D- Multi layer plot
if plot==2:
    nds_frontX=np.array(nds_frontX)
    nds_frontY=np.array(nds_frontY)
    nds_frontZ=np.array(nds_frontZ)
    pressures=np.array(pressures)

    x = nds_frontX.reshape(12, 4)
    y = nds_frontY.reshape(12, 4)
    z = nds_frontZ.reshape(12, 4)
    p =pressures.reshape(12,4)

    print(x[1,:])
    fig = plt.figure(figsize=plt.figaspect(2)*1)
    ax = plt.axes(projection='3d')
    ax.contourf(x, y, p, 48, zdir='z', offset=0)
    ax.contourf(x, y, p, 48, zdir='z', offset=5)
    ax.contourf(x, y, p, 48, zdir='z', offset=10)

    ax.set_title('Test')
    plt.triplot(tri_front, '-k')
    plt.show()

######################################## BACK up ########################################
#triangles_front= [
#    [0, 1, 4],
#    [1, 2, 5],
#    [2, 3, 6],
#    [0, 4, 8],
#    [4, 1, 9],
#    [1, 5, 9],
#    [5, 2, 10],
#    [2, 6, 10],
#    [6, 3, 11],
#    [3, 7, 11],
#    [8, 4, 12],
#    [4, 9, 12],
#    [9, 5, 13],
#    [5, 10, 13],
#    [10, 6, 14],
#    [6, 11, 14],
#    [11, 7, 15],
#    [8, 12, 16],
#    [12, 9, 17],
#    [9, 13, 17],
#    [13, 10, 18],
#    [10, 14, 18],
#    [14, 11, 19],
#    [11, 15, 19],
#    [16, 12, 20],
#    [12, 17, 20],
#    [17, 13, 21],
#    [13, 18, 21],
#    [18, 14, 22],
#    [14, 19, 22],
#    [19, 15, 23],
#    [16, 20, 24],
#    [20, 17, 25],
#    [17, 21, 25],
#    [21, 18, 26],
#    [18, 22, 26],
#    [22, 19, 27],
#    [19, 23, 27],
#    [24, 20, 28],
#    [20, 25, 28],
#    [25, 21, 29],
#    [21, 26, 29],
#    [26, 22, 30],
#    [22, 27, 30],
#    [27, 23, 31],
#    [24, 28, 32],
#    [28, 25, 33],
#    [25, 29, 33],
#    [29, 26, 34],
#    [26, 30, 34],
#    [30, 27, 35],
#    [27, 31, 35],
#    [32, 28, 36],
#    [28, 33, 36],
#    [33, 29, 37],
#    [29, 34, 37],
#    [34, 30, 38],
#    [30, 35, 38],
#    [35, 31, 39],
#    [32, 36, 40],
#    [36, 33, 41],
#    [33, 37, 41],
#    [37, 34, 42],
#    [34, 38, 42],
#    [38, 35, 43],
#    [35, 39, 43],
#    [40, 36, 44],
#    [36, 41, 44],
#    [41, 37, 45],
#    [37, 42, 45],
#    [42, 38, 46],
#    [38, 43, 46],
#    [43, 39, 47],
#    [44, 41, 45],
#    [45, 42, 46],
#    [46, 43, 47]
#]
##2D-Contour Plot
#if plot==1:
#    plt.triplot(tri_front, '-k')
#    plt.tricontourf(tri_front, pressures)
#    plt.colorbar()
#    plt.show()

    #fig,ax1 = plt.subplots()
    #ax1.set_title('Pressure plot - Artificial Skin Array (Dorsal). Timestep: %i ms' % timeStep)
    #ax1.set_xlabel('Contra-lateral [mm]',fontsize=12)
    #ax1.set_ylabel('Inferior [mm]',fontsize=12)
    #ax1.invert_yaxis()
    #plt.triplot(tri_front, '-k')
    #tcf=plt.tricontourf(tri_front, pressure_front[:,0],levels,cmap='jet')    
    #cbar=plt.colorbar(ticks=level_labels,cmap='jet')
    #cbar.ax.set_yticklabels(["%1.1f" % y for y in level_labels], fontsize=12)
    #cbar.set_label('Force in [N]', labelpad=40, y=0.5, rotation=270,fontsize=12)
    ## Add label shwoing the time to the plot
    #style = dict(size=12, color='black')
    #t1=ax1.text(97, 90,'t= 0 ms', **style) #add annotation showing the time
    #plt.show()

    #for k in range(1,num_measurements,plotStep): #num_measurements
    #    #update=pressure_front[k]
    #    print("k")
    #    print(k)
    #    plt.tricontourf(tri_front, pressure_front[:,k],levels,cmap='jet')
    #    #ax1.text(95, 90,'t= %i ms' % time[k], **style) #add annotation showing the time
    #    t1.set_text('t= %1.2f secs' % (time[k]))
    #    print(pressure_front[17,k])
    #    plt.draw()  
    #    plt.pause(1.0)
    #    #fig.clear()
