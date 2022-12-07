import math
import os
from enum import Enum
import numpy as np
import openpyxl
from matplotlib import pyplot as plt

from plotter import Plot


class Label(Enum):

    NO_ORTHOSIS_0 = 0
    NO_ORTHOSIS_45 = 1
    NO_ORTHOSIS_90 = 2
    CORRECT_ORTHOSIS_0 = 3
    CORRECT_ORTHOSIS_45 = 4
    CORRECT_ORTHOSIS_90 = 5
    UP_ORTHOSIS_0 = 6
    UP_ORTHOSIS_45 = 7
    UP_ORTHOSIS_90 = 8
    DOWN_ORTHOSIS_0 = 9
    DOWN_ORTHOSIS_45 = 10
    DOWN_ORTHOSIS_90 = 11
    CLOCK_ROT_ORTHOSIS_0 = 12
    CLOCK_ROT_ORTHOSIS_45 = 13
    CLOCK_ROT_ORTHOSIS_90 = 14
    COUTNERCLOCK_ROT_ORTHOSIS_0 = 15
    COUTNERCLOCK_ROT_ORTHOSIS_45 = 16
    COUTNERCLOCK_ROT_ORTHOSIS_90 = 16



class LabelDataAnalysis:


    def __init__(self, data_folder, label):

        self.data_folder = data_folder
        self.dir = os.listdir(self.data_folder)
        self.label = label



    def mean_variance_at_degrees(self,at_arm_degrees, linearized = True):


        p_sensor_values = []
        rot_sensor_values = []
        for file in self.dir:
            if (linearized and "LIN" not in file) or (not linearized and "RAW" not in file): continue

            path = self.data_folder + "/" + file
            sample = SampleDataAnalysis(self.label, file_path=path)

            p_sensors, rot_sensors, time_stamp = sample.sensor_values_at_arm_degrees(at_arm_degrees)
            p_sensor_values.append(p_sensors)
            rot_sensor_values.append(rot_sensors)


        ## MEAN
        p_sensor_avg = np.zeros(len(p_sensor_values[0]))
        for sensor_values in p_sensor_values:
            p_sensor_avg += sensor_values

        p_sensor_avg = p_sensor_avg/len(p_sensor_values)

        rot_sensor_avg = np.zeros(len(rot_sensor_values[0]))
        for sensor_values in  rot_sensor_values:
            rot_sensor_avg += sensor_values
        rot_sensor_avg = rot_sensor_avg/len(rot_sensor_values)
        ## Variance

        p_sensor_var = np.zeros(len(p_sensor_values[0]))
        for sensor_values in p_sensor_values:
            p_sensor_var += (p_sensor_avg - sensor_values)**2

        p_sensor_var = p_sensor_var/len(p_sensor_values)

        rot_sensor_var = np.zeros(len(rot_sensor_values[0]))
        for sensor_values in rot_sensor_values:
            rot_sensor_var += (rot_sensor_avg - sensor_values)**2
        rot_sensor_var = rot_sensor_var/len(rot_sensor_values)

        return p_sensor_avg, p_sensor_var, rot_sensor_avg, rot_sensor_var




class SampleDataAnalysis:

    ## at index -2 is arm degree, at index -1 is orthosis degree

    ## either initialize from csv file, or from prebuilt arrays
    def __init__(self, label, file_path=None, p_sensors=None, rot_sensors=None, sample_time_stamps=None):

        self.label = label

        if file_path is not None:
            self.file_path = file_path
            self.sheet = openpyxl.load_workbook(self.file_path).active
            self.p_sensors, self.rot_sensors, self.sample_time_stamps = self.load_sample()
        if p_sensors is not None and sample_time_stamps is not None and rot_sensors is not None:
            self.p_sensors = p_sensors
            self.rot_sensors = rot_sensors
            self.sample_time_stamps = sample_time_stamps

        self.time_step = self.sample_time_stamps[1] - self.sample_time_stamps[0]
        self.num_p_sensors = len(self.p_sensors[0])
        self.num_rot_sensors = len(self.rot_sensors[0])
        self.num_time_steps = len(self.sample_time_stamps)

    def load_sample(self):

        num_sensors = self.sheet.max_row
        num_columns = self.sheet.max_column

        sample_time_stamps = np.zeros(num_columns)
        p_sensors = np.zeros((num_columns,num_sensors - 6))
        rot_sensors = np.zeros((num_sensors,2))
        for i in range(num_columns):
            sample_time_stamps[i] = self.sheet.cell(2,i+1).value
            rot_sensors[i][0] = self.sheet.cell(num_sensors-1,i+1).value
            rot_sensors[i][1] = self.sheet.cell(num_sensors,i+1).value

            for j in range(0,num_sensors - 6):
                p_sensors[i][j] = self.sheet.cell(j + 4,i + 1).value

        return p_sensors, rot_sensors, sample_time_stamps

    def sensor_values_at_arm_degrees(self, at_arm_degrees):

        min_difference = abs(at_arm_degrees - self.rot_sensors[0][0])
        index = 0

        for i in range(1,self.num_time_steps):
            difference = abs(at_arm_degrees - self.rot_sensors[i][0])
            if difference < min_difference:
                index = i
                min_difference = difference
            if difference < 10e-5:
                break

        return self.p_sensors[index], self.rot_sensors[index], self.sample_time_stamps[index]

    def sensorwise_avg(self, avg_range = None):
        if avg_range is None: avg_range = range(self.num_time_steps)
        else: avg_range = range(max(0,avg_range[0]),min(self.num_time_steps,avg_range[-1]+1))

        p_sensors_avg = np.zeros(self.num_p_sensors)
        rot_sensors_avg = np.zeros(2)
        for i in avg_range:
            p_sensors_avg += self.p_sensors[i]
            rot_sensors_avg += self.rot_sensors[i]


        return p_sensors_avg/len(avg_range), rot_sensors_avg/len(avg_range)

    def timewise_avg(self):

        p_time_stamp_avg = np.zeros(self.num_time_steps)
        rot_time_stamp_avg = np.zeros(self.num_time_steps)
        for i in range(self.num_time_steps):
            p_time_stamp_avg[i] = sum(self.p_sensors[i])/self.num_p_sensors
            rot_time_stamp_avg[i] = sum(self.rot_sensors[i])/self.num_rot_sensors

        return p_time_stamp_avg,rot_time_stamp_avg

    def total_avg(self):
        p_avg,rot_avg = self.sensorwise_avg(range(0,self.num_time_steps))
        return sum(p_avg)/len(p_avg), sum(rot_avg)/len(rot_avg)

    def avg_at_arm_degrees(self,at_arm_degrees):

        p_values,rot_values,time_stamp = self.sensor_values_at_arm_degrees(at_arm_degrees)

        return sum(p_values)/len(p_values), sum(rot_values)/len(rot_values)

    def get_p_value(self,column, row):
        return self.p_sensors[column][row]
    def get_rot_values(self,column,row):
        return self.rot_sensors[column][row]
    def get_p_values_at_time(self,column):
        return self.p_sensors[column]
    def get_rot_values_at_time(self,column):
        return self.rot_sensors[column]
    def get_time_at(self,column):
        return self.sample_time_stamps[column]
    def get_sensor_values(self,sensor):
        return self.p_sensors[:,sensor]




    ## makes sample smaller in time domain, sets k time stamps, and for each computes the average of the timestamps that fall into the new time stamp range
    def shrink_sample_time_domain(self,k_shrink):

        shrink_ratio = math.ceil(self.num_time_steps/k_shrink)

        new_p_sensors = np.zeros((k_shrink,self.num_p_sensors))
        new_rot_sensors = np.zeros((k_shrink,self.num_rot_sensors))
        new_sample_time_stamps = np.zeros(k_shrink)
        for i in range(k_shrink):
            ## Note if self.num_time_steps is not divisible by k_shrink, the last
            new_p_sensors[i], new_rot_sensors[i] = self.sensorwise_avg(range(i*shrink_ratio,(i+1)*shrink_ratio))
            new_sample_time_stamps[i] = self.sample_time_stamps[i*shrink_ratio]

        return SampleDataAnalysis(p_sensors=new_p_sensors,rot_sensors=new_rot_sensors,sample_time_stamps=new_sample_time_stamps,label=self.label)

    def extrema_pressure_time_stamp(self, in_range=None):
        if in_range is None: in_range = range(self.num_time_steps)

        max_i = in_range[0]
        max_press = sum(self.p_sensors[in_range[0]])
        min_i = in_range[0]
        min_press = sum(self.p_sensors[in_range[0]])


        for i in in_range:

            p_sum = sum(self.p_sensors[i])
            if p_sum < min_press:
                min_i = i
                min_press = p_sum
            if p_sum > max_press:
                max_i = i
                max_press = p_sum

        return self.p_sensors[min_i], self.p_sensors[max_i]

    def extrema_pressure_sensor(self):

        min_i = 0
        min_press = sum(self.p_sensors[:,0])
        max_i = 0
        max_press = sum(self.p_sensors[:,0])

        for i in range(1,self.num_p_sensors):
            p_sum = sum(self.p_sensors[:,i])
            if p_sum < min_press:
                min_i = i
                min_press = p_sum
            if p_sum > max_press:
                max_i = i
                max_press = p_sum

        return self.p_sensors[:,min_i], self.p_sensors[:,max_i]

    def copy(self):

        return SampleDataAnalysis(self.label,p_sensors=np.copy(self.p_sensors),rot_sensors=np.copy(self.rot_sensors),sample_time_stamps=np.copy(self.sample_time_stamps))


def plot_pressure_data(sensor_array):
    plot = Plot(10)
    plt.pause(0.05)
    plot.update_pressure_plot(sensor_array[1:49], sensor_array[49:71], sensor_array[71:])
    plt.pause(20)


def plot_angle_data(angle):
    # angle[0] arm
    # angle[1] orthosis
    pass


# da = LabelDataAnalysis("C:/Users/majag/Desktop/marble/data/no_orthosis_90", Label.NO_ORTHOSIS_90)
# p_sensor_avg, p_sensor_var, rot_sensor_avg, rot_sensor_var = da.mean_variance_at_degrees(90)
# plot_pressure_data(p_sensor_avg)
# plot_angle_data(rot_sensor_avg)
# print(da.mean_variance_at_degrees(90))

# add to the plot circle with radius of the value of variance
# both avg and variance